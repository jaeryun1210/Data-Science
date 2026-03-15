import os
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Alignment

def get_last_week_dates():
    """실행일 기준 지난주 월~일 날짜 추출"""
    today = datetime.now()
    days_to_last_monday = today.weekday() + 7
    last_monday = today - timedelta(days=days_to_last_monday)
    
    dates = []
    for i in range(7):
        dt = last_monday + timedelta(days=i)
        dates.append(f"{dt.year}년 {dt.month}월 {dt.day}일")
    return dates

def process_workout_auth(txt_file, excel_file):
    try:
        wb = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        print(f"❌ '{excel_file}' 파일이 없습니다. 먼저 양식을 만들어주세요.")
        return
        
    ws = wb.active
    
    # 1. 엑셀 명단 기반 매핑 사전 생성 (동명이인 없음 전제)
    # 엑셀에 '홍길동'이 있으면 {'홍길동': '홍길동', '길동': '홍길동'} 생성
    name_map = {}
    for row in range(2, ws.max_row + 1):
        full_name = ws.cell(row=row, column=1).value
        if full_name and isinstance(full_name, str):
            full_name = full_name.strip()
            name_map[full_name] = full_name # 풀네임 매핑
            if len(full_name) >= 3:
                name_map[full_name[1:]] = full_name # 성 뺀 이름 매핑

    # 2. 카톡 텍스트 파싱 준비
    target_dates = get_last_week_dates()
    date_pattern = re.compile(r'-+ (\d{4}년 \d{1,2}월 \d{1,2}일) .+-+')
    msg_pattern = re.compile(r'\[(.+?)\] \[(.+?)\] (.+)')
    
    # 한글(이름) + 공백(선택) + 숫자(회차) 패턴
    auth_pattern = re.compile(r'([가-힣]{2,4})\s*(\d+)')

    auth_data = {} 
    current_date = None

    # 3. 대화 내용 분석
    with open(txt_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            
            # 날짜 확인
            date_match = date_pattern.match(line)
            if date_match:
                current_date = date_match.group(1)
                continue
            
            # 지난주 데이터만 처리
            if current_date not in target_dates:
                continue
                
            # 메시지 확인
            msg_match = msg_pattern.match(line)
            if msg_match:
                content = msg_match.group(3).strip()
                
                # 메시지 내에서 인증 패턴 찾기
                # search를 쓰면 "오늘도 완료 길동 1" 같은 문장도 인식합니다.
                # 오직 "길동 1" 처럼 딱 떨어지는 것만 잡으려면 auth_pattern.fullmatch(content)로 변경하세요.
                auth_match = auth_pattern.search(content)

                if auth_match:
                    input_name = auth_match.group(1) # 입력된 이름 (예: 길동)
                    
                    # 입력된 이름이 엑셀 명단(name_map)에 존재하는 경우만 인정
                    if input_name in name_map:
                        real_name = name_map[input_name] # 무조건 엑셀 풀네임(홍길동)으로 치환
                        
                        if real_name not in auth_data:
                            auth_data[real_name] = set()
                        # 중복 인증 방지를 위해 '날짜'를 set에 저장 (하루 1회만 인정)
                        auth_data[real_name].add(current_date)

    # 4. 엑셀에 결과 그리기
    fill_style = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # 연노랑 배경
    align_style = Alignment(horizontal='center', vertical='center') # 가운데 정렬

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        target_count = ws.cell(row=row, column=2).value
        
        if not name or not isinstance(target_count, int):
            continue
            
        name = name.strip()
        actual_count = len(auth_data.get(name, set())) # 실제 인증한 날짜 수
        
        start_col = 3 # C열부터 시작
        
        # 목표치만큼 배경색 칠하기 (기존 데이터 초기화 포함)
        for i in range(target_count):
            cell = ws.cell(row=row, column=start_col + i)
            cell.fill = fill_style
            cell.value = "" 
            
        # 실제 인증한 횟수만큼 'O' 표시
        for i in range(actual_count):
            cell = ws.cell(row=row, column=start_col + i)
            cell.value = 'O'
            cell.alignment = align_style

    # 5. 저장
    base_dir = os.path.dirname(__file__)
    result_file = os.path.join(base_dir, '미모인증_이번주결과.xlsx')
    wb.save(result_file)
    print(f"✅ 정산 기간: {target_dates[0]} ~ {target_dates[-1]}")
    print(f"✅ '{result_file}' 파일 생성 완료!")

# --- 실행 ---
# 파일 이름이 맞는지 확인하세요!
process_workout_auth('연지/xxxcn/mm.txt', '연지/xxxcn/미라클모닝.xlsx')